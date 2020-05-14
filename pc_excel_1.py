import openpyxl 
# 利用openpyxl.Workbook()函数创建新的workbook（工作簿）对象，就是创建新的空的Excel文件。
wb=openpyxl.Workbook()
# wb.active就是获取这个工作簿的活动表，通常就是第一个工作表。
sheet=wb.active
#可以用.title给工作表重命名。现在第一个工作表的名称就会由原来默认的“sheet1”改为"new title"
sheet.title='我的PyExcel'
# 把'漫威宇宙'赋值给第一个工作表的A1单元格，就是往A1的单元格中写入了'漫威宇宙'
sheet['A1'] = 'Excel'
rows=[['e1','e2','e3'],['x1','x2','x3']]
for i in rows:
    sheet.append(i)
print(rows)
wb.save('myexcel.xlsx')

# 读取的代码：
wb = openpyxl.load_workbook('myexcel.xlsx')
sheet = wb['我的PyExcel']
sheetname = wb.sheetnames
print(sheetname)
A1_cell = sheet['A1']
A1_value = A1_cell.value
print(sheet['A2'].value)